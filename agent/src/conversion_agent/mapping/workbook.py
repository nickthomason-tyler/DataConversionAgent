"""Parser for tool-generated lookup crosswalk workbooks."""

from __future__ import annotations

import json

import openpyxl

from conversion_agent.core.errors import WorkbookError

from .model import CrosswalkWorkbook, Section, SourceRow


def _parse_lookup_spec(wb) -> dict:
    try:
        ws = wb["LookupSpec"]
    except KeyError as exc:
        raise WorkbookError("Workbook has no LookupSpec worksheet") from exc
    rows = list(ws.iter_rows(values_only=True))
    for r in rows:
        if r and r[0] == "spec":
            try:
                spec = json.loads("".join(c for c in r[4:] if isinstance(c, str)))
            except json.JSONDecodeError as exc:
                raise WorkbookError(f"Invalid LookupSpec JSON: {exc}") from exc
            if not isinstance(spec, dict):
                raise WorkbookError("LookupSpec JSON must be an object")
            return spec
    return {}


def _spec_counts(spec: dict) -> dict[str, dict[str, tuple[int, int]]]:
    """Validate the generator contract shape while deriving visible column counts."""
    modules = spec.get("modules", {})
    if not isinstance(modules, dict):
        raise WorkbookError("LookupSpec modules must be an object")

    counts_by_tab: dict[str, dict[str, tuple[int, int]]] = {}
    for module, module_spec in modules.items():
        if not isinstance(module, str) or not isinstance(module_spec, dict):
            raise WorkbookError("LookupSpec module entries must be objects")
        type_queries = module_spec.get("typeQueries", {})
        if not isinstance(type_queries, dict):
            raise WorkbookError(f"LookupSpec module {module} has invalid typeQueries")
        for section, section_spec in type_queries.items():
            if not isinstance(section, str) or not isinstance(section_spec, dict):
                raise WorkbookError(f"LookupSpec module {module} has an invalid section")
            source = section_spec.get("source")
            destination = section_spec.get("destination")
            if not isinstance(source, dict) or not isinstance(destination, dict):
                raise WorkbookError(f"LookupSpec section {section} has invalid endpoints")
            source_columns = source.get("columns")
            destination_columns = destination.get("columns")
            if not isinstance(source_columns, list) or not isinstance(destination_columns, list):
                raise WorkbookError(f"LookupSpec section {section} has invalid columns")
            counts_by_tab.setdefault(module, {})[section] = (
                len(source_columns),
                len(destination_columns),
            )
    return counts_by_tab


def _parse_visible_tab(ws, tab: str, spec_counts: dict[str, tuple[int, int]]) -> list[Section]:
    rows = list(ws.iter_rows(values_only=True))
    sections: list[Section] = []
    i = 0
    while i < len(rows):
        r = rows[i] or ()
        if "Source DB" in r:
            title = None
            for back in range(i - 1, max(i - 3, -1), -1):
                if rows[back] and rows[back][0]:
                    title = str(rows[back][0]).strip()
                    break
            # Column layout is contiguous: source cols, then dest cols, then
            # notes. Prefer the LookupSpec's column counts — merged header
            # cells lose their duplicated labels after a round-trip through
            # openpyxl, so counting "Source DB" cells undercounts.
            if title in spec_counts:
                n_src, n_dst = spec_counts[title]
                src_cols = list(range(1, n_src + 1))
                dst_cols = list(range(n_src + 1, n_src + n_dst + 1))
            else:
                src_cols = [c + 1 for c, v in enumerate(r) if v == "Source DB"]
                dst_cols = [c + 1 for c, v in enumerate(r) if v == "Destination DB"]
            header = rows[i + 1] if i + 1 < len(rows) else ()
            notes_col = None
            for c, v in enumerate(header):
                if v and "note" in str(v).lower():
                    notes_col = c + 1
            sec = Section(
                tab=tab,
                title=title or "?",
                src_cols=src_cols,
                dst_cols=dst_cols,
                notes_col=notes_col,
                header_row=i + 2,
            )
            j = i + 2
            while j < len(rows) and "Source DB" not in (rows[j] or ()):
                j += 1
            data_end = j - 1 if j < len(rows) else j
            for k in range(i + 2, data_end):
                dr = rows[k] or ()
                src = tuple(
                    str(dr[c - 1]).strip() if c - 1 < len(dr) and dr[c - 1] is not None else ""
                    for c in src_cols
                )
                if not any(src):
                    continue
                dst = tuple(
                    str(dr[c - 1]).strip() if c - 1 < len(dr) and dr[c - 1] is not None else ""
                    for c in dst_cols
                )
                sec.rows.append(SourceRow(row_idx=k + 1, values=src, existing=dst))
            if sec.rows:
                sections.append(sec)
            i = j
        else:
            i += 1
    return sections


def _parse_hidden_tab(ws) -> dict[str, tuple[list[list[str]], dict[str, list[str]]]]:
    """Return {section title: (dest value lists, cascade map)}."""
    out: dict[str, tuple[list[list[str]], dict[str, list[str]]]] = {}
    current: str | None = None
    for row_number, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if not r or r[0] is None:
            continue
        if len(r) < 3:
            raise WorkbookError(f"Malformed hidden sheet row {row_number}")
        first = str(r[0])
        if isinstance(r[1], (int, float)) and isinstance(r[2], (int, float)):
            current = first.strip()
            out[current] = ([], {})
        elif current and ".head" in first:
            values = [str(v) for v in r[3:] if v is not None]
            out[current][0].append(values)
        elif current and len(r) > 2 and isinstance(r[2], (int, float)):
            values = [str(v) for v in r[3:] if v is not None]
            out[current][1][first] = values
    return out


def load(path: str) -> CrosswalkWorkbook:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    model = CrosswalkWorkbook(path=path, spec=_parse_lookup_spec(wb))
    counts_by_tab = _spec_counts(model.spec)
    hidden: dict[str, dict] = {}
    for name in wb.sheetnames:
        if name.endswith("Hidden"):
            hidden[name[: -len(" Hidden")]] = _parse_hidden_tab(wb[name])
    for name in wb.sheetnames:
        if name.endswith("Hidden") or name == "LookupSpec":
            continue
        for sec in _parse_visible_tab(wb[name], name, counts_by_tab.get(name, {})):
            lists = hidden.get(name, {}).get(sec.title)
            if lists:
                sec.dest_lists, sec.cascade = lists
            model.sections.append(sec)
    wb.close()
    return model
