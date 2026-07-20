"""Write proposals back into a copy of the client's workbook.

The output is the same workbook the team already reviews in Excel: destination
cells filled with byte-exact pick-list values, the notes column carrying
method + confidence + rationale, and a fill color per lane so reviewers can
scan for what needs attention. Human-entered destinations are never touched.
"""

from __future__ import annotations

import openpyxl
from openpyxl.styles import PatternFill

from .model import CrosswalkWorkbook

FILL_AUTO = PatternFill("solid", start_color="C6EFCE")      # green — deterministic
FILL_PROPOSED = PatternFill("solid", start_color="FFEB9C")  # yellow — model-proposed


def write(model: CrosswalkWorkbook, out_path: str) -> dict:
    wb = openpyxl.load_workbook(model.path)
    written = {"auto": 0, "llm": 0}
    for sec in model.sections:
        ws = wb[sec.tab]
        for row_idx, prop in sec.proposals.items():
            fill = FILL_AUTO if prop.method != "llm" else FILL_PROPOSED
            for col, value in zip(sec.dst_cols, prop.dest):
                cell = ws.cell(row=row_idx, column=col)
                if cell.value not in (None, ""):
                    continue  # safety: never overwrite
                cell.value = value
                cell.fill = fill
            if sec.notes_col:
                note_cell = ws.cell(row=row_idx, column=sec.notes_col)
                if note_cell.value in (None, ""):
                    note_cell.value = prop.note
            written["llm" if prop.method == "llm" else "auto"] += 1
    wb.save(out_path)
    return written
