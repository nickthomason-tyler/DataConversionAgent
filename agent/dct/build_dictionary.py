"""Generate the DCT data dictionary from Tyler's DCT-DB release artifacts.

Usage:
    python build_dictionary.py <DCT-DB_datatype_operation_values_*.xlsx> \
        <DCT-DB_V*.docx> <version-label>

Reads the datatype/values spreadsheet (authoritative table+column schema) and
the DCT-DB Word documentation (module groupings and table descriptions), then
writes the canonical packaged dictionary resource at
``src/conversion_agent/resources/data/dct/dictionary.yaml``. Re-run per DCT
release and review the generated resource before committing it.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
import yaml
from docx import Document

# docx heading -> module key
MODULE_HEADINGS = {
    "Contact": "contacts",
    "Professional Licensing": "professional_license",
    "Business": "business",
    "Business Licensing": "business_license",
    "Environmental Health": "environmental_health",
    "Code Enforcement": "code_enforcement",
    "Permit": "permits",
    "Plans": "plans",
    "Projects": "projects",
    "Inspection Case": "inspection_case",
    "Inspections": "inspections",
    "Financial Tables": "finance",
    "Parcel": "parcel",
    "Bonds": "bonds",
    "Impact": "impact",
    "Submittals": "submittals",
    "Meetings & Hearing": "meetings_hearings",
    "Objects": "objects",
    "Time Tracking": "time_tracking",
    "Custom Fields / Additional Info": "custom_fields",
    "Attachment Documents": "attachments",
    "Parent Record Associations": "parent_associations",
    "Tax Remittance": "tax_remittance",
}


def read_schema(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    tables: dict[str, dict] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        table, column, dtype, max_len, nullable = r[0], r[1], r[2], r[3], r[4]
        t = tables.setdefault(str(table), {"columns": {}})
        t["columns"][str(column)] = {
            "type": str(dtype),
            "max_length": int(max_len) if isinstance(max_len, (int, float)) else None,
            "nullable": bool(int(nullable)) if isinstance(nullable, (int, float)) else None,
        }
    return tables


def read_docs(docx_path: Path, table_names: set[str]) -> tuple[dict, dict, dict]:
    """Return (module_of_table, table_descriptions, module_notes)."""
    doc = Document(docx_path)
    module_of, descriptions, module_notes = {}, {}, {}
    # longest-first so business_license_contact matches before business
    names_sorted = sorted(table_names, key=len, reverse=True)
    current = None
    for p in doc.paragraphs:
        text = p.text.strip()
        if not text:
            continue
        if p.style.name.startswith("Heading"):
            current = MODULE_HEADINGS.get(text.rstrip(": ").strip())
            continue
        if current is None:
            continue
        module_notes.setdefault(current, text if ":" not in text[:40] else "")
        # lines usually read "table_name: description" (or "table_name o desc")
        lowered = text.lower()
        for name in names_sorted:
            if lowered.startswith(name):
                rest = text[len(name) :].lstrip(" :o-–").strip()
                module_of.setdefault(name, current)
                if rest and name not in descriptions:
                    descriptions[name] = re.sub(r"\s+", " ", rest)[:300]
                break
    return module_of, descriptions, module_notes


def main() -> None:
    xlsx_path, docx_path, version = Path(sys.argv[1]), Path(sys.argv[2]), sys.argv[3]
    tables = read_schema(xlsx_path)
    module_of, descriptions, _ = read_docs(docx_path, set(tables))

    # Fallbacks for tables the docx doesn't name individually: inherit the
    # module of the longest assigned table name they extend (e.g.
    # permit_activity_additional_fields -> permits), then fixed buckets.
    assigned_sorted = sorted(module_of, key=len, reverse=True)
    for name in tables:
        if name in module_of:
            continue
        lname = name.lower()
        parent = next((a for a in assigned_sorted if lname.startswith(a.lower())), None)
        if parent:
            module_of[name] = module_of[parent]
        elif "custom_field" in lname:
            module_of[name] = "custom_fields"
        elif lname.startswith("vw_"):
            module_of[name] = "views"
        elif lname.startswith(("error", "step", "filestream", "filetable")):
            module_of[name] = "internal"

    for name, entry in tables.items():
        entry["module"] = module_of.get(name, "unassigned")
        if name in descriptions:
            entry["description"] = descriptions[name]

    out = {
        "version": version,
        "source": {
            "schema": xlsx_path.name,
            "documentation": docx_path.name,
        },
        "table_count": len(tables),
        "column_count": sum(len(t["columns"]) for t in tables.values()),
        "tables": tables,
    }
    dest = (
        Path(__file__).resolve().parents[1]
        / "src/conversion_agent/resources/data/dct/dictionary.yaml"
    )
    dest.write_text(yaml.safe_dump(out, sort_keys=True, width=100))
    print(f"wrote {dest}: {out['table_count']} tables, {out['column_count']} columns")


if __name__ == "__main__":
    main()
