from __future__ import annotations

import shutil
import zipfile
from pathlib import Path

import openpyxl
import pytest
from lxml import etree
from openpyxl.workbook.defined_name import DefinedName

from conversion_agent.core.errors import OutputError
from conversion_agent.mapping import writeback
from conversion_agent.mapping.model import CrosswalkWorkbook, Proposal, Section, SourceRow


def _crosswalk(
    tmp_path: Path,
    *,
    existing_note: str = "",
    proposal_dest: tuple[str, ...] = ("Approved ",),
    proposal_note: str = "verified ",
) -> tuple[Path, CrosswalkWorkbook]:
    path = tmp_path / "crosswalk.xlsx"
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Permits"
    sheet["A1"] = "Type"
    sheet["A3"] = "Legacy type"
    sheet["B3"] = None
    sheet["C3"] = existing_note or None
    hidden = book.create_sheet("Permits Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "keep hidden"
    lookup = book.create_sheet("LookupSpec")
    lookup["A1"] = "spec"
    book.defined_names.add(DefinedName("KeepMe", attr_text="'Permits'!$A$3"))
    book.save(path)
    book.close()

    # openpyxl cannot retain x14 extensions itself.  Add a minimal extension
    # directly to the synthetic package so preservation is exercised here.
    staged = tmp_path / "staged.xlsx"
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(staged, "w") as target:
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                root = etree.fromstring(content)
                ext = etree.fromstring(
                    b'<extLst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                    b'<ext uri="{CCE6A557-97BC-4B89-ADB6-D9C93CAAB3DF}">'
                    b'<x14:dataValidations xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"/>'
                    b'</ext></extLst>'
                )
                root.append(ext)
                content = etree.tostring(root, xml_declaration=True, encoding="UTF-8")
            target.writestr(item, content)
    shutil.move(staged, path)

    section = Section(
        tab="Permits",
        title="Type",
        src_cols=[1],
        dst_cols=[2],
        notes_col=3,
        header_row=2,
        rows=[SourceRow(row_idx=3, values=("Legacy type",), existing=("",))],
        proposals={
            3: Proposal(
                dest=proposal_dest, method="exact", confidence=1.0, note=proposal_note
            )
        },
    )
    return path, CrosswalkWorkbook(path=str(path), spec={"modules": {}}, sections=[section])


def _rewrite_package(path: Path, mutate) -> None:
    staged = path.with_suffix(".staged.xlsx")
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(staged, "w") as target:
        for item in source.infolist():
            replacement = mutate(item.filename, source.read(item.filename))
            if replacement is not None:
                target.writestr(item, replacement)
    shutil.move(staged, path)


def _cell_text(path: Path, ref: str) -> str:
    with zipfile.ZipFile(path) as package:
        root = etree.fromstring(package.read("xl/worksheets/sheet1.xml"))
    return root.xpath(
        f"string(.//m:c[@r='{ref}']/m:is/m:t)", namespaces={"m": writeback.NS_MAIN}
    )


def test_write_refuses_input_as_output(tmp_path: Path) -> None:
    path, model = _crosswalk(tmp_path)

    with pytest.raises(OutputError, match="input workbook"):
        writeback.write(model, str(path))


def test_failed_verification_keeps_existing_output(monkeypatch, tmp_path: Path) -> None:
    _, model = _crosswalk(tmp_path)
    output = tmp_path / "out.xlsx"
    output.write_bytes(b"existing")
    monkeypatch.setattr(
        writeback,
        "verify_output",
        lambda *args, **kwargs: (_ for _ in ()).throw(OutputError("bad workbook")),
    )

    with pytest.raises(OutputError, match="bad workbook"):
        writeback.write(model, str(output))

    assert output.read_bytes() == b"existing"
    assert not list(tmp_path.glob(".out.xlsx.*.tmp"))


def test_write_preserves_human_no_good_match_note_without_overwrite(tmp_path: Path) -> None:
    _, model = _crosswalk(
        tmp_path,
        existing_note="human review",
        proposal_dest=("",),
        proposal_note="NO GOOD MATCH: needs review",
    )
    output = tmp_path / "out.xlsx"

    report = writeback.write(model, str(output))

    assert _cell_text(output, "C3") == "human review"
    assert report.as_dict() == {"auto": 0, "llm": 0}
    assert report.note_cells == 0


def test_write_replaces_human_no_good_match_note_with_overwrite(tmp_path: Path) -> None:
    _, model = _crosswalk(
        tmp_path,
        existing_note="human review",
        proposal_dest=("",),
        proposal_note="NO GOOD MATCH: needs review",
    )
    output = tmp_path / "out.xlsx"

    report = writeback.write(model, str(output), overwrite=True)

    assert _cell_text(output, "C3") == "NO GOOD MATCH: needs review"
    assert report.as_dict() == {"auto": 1, "llm": 0}
    assert report.note_cells == 1


def test_write_refuses_hard_link_alias_of_input(tmp_path: Path) -> None:
    source, model = _crosswalk(tmp_path)
    alias = tmp_path / "input-alias.xlsx"
    alias.hardlink_to(source)
    before = source.read_bytes()

    with pytest.raises(OutputError, match="input workbook"):
        writeback.write(model, str(alias))

    assert source.read_bytes() == before
    assert alias.read_bytes() == before


def test_write_preserves_protected_parts_extensions_and_trailing_spaces(tmp_path: Path) -> None:
    source, model = _crosswalk(tmp_path)
    output = tmp_path / "out.xlsx"

    report = writeback.write(model, str(output))

    assert report.deterministic_rows == 1
    assert report.model_rows == 0
    assert report.destination_cells == 1
    assert report.note_cells == 1
    assert report.as_dict() == {"auto": 1, "llm": 0}
    assert report["auto"] == 1
    assert report["llm"] == 0
    with zipfile.ZipFile(source) as source_zip, zipfile.ZipFile(output) as output_zip:
        assert source_zip.read("xl/workbook.xml") == output_zip.read("xl/workbook.xml")
        assert source_zip.read("xl/_rels/workbook.xml.rels") == output_zip.read(
            "xl/_rels/workbook.xml.rels"
        )
        assert source_zip.read("xl/worksheets/sheet2.xml") == output_zip.read(
            "xl/worksheets/sheet2.xml"
        )
        source_root = etree.fromstring(source_zip.read("xl/worksheets/sheet1.xml"))
        output_root = etree.fromstring(output_zip.read("xl/worksheets/sheet1.xml"))
        ns = {"m": writeback.NS_MAIN}
        assert etree.tostring(source_root.find("m:extLst", ns)) == etree.tostring(
            output_root.find("m:extLst", ns)
        )
        assert output_root.xpath("string(.//m:c[@r='B3']/m:is/m:t)", namespaces=ns) == "Approved "
        assert output_root.xpath("string(.//m:c[@r='C3']/m:is/m:t)", namespaces=ns) == "verified "


def test_write_reports_style_failures_without_losing_value(monkeypatch, tmp_path: Path) -> None:
    _, model = _crosswalk(tmp_path)
    output = tmp_path / "out.xlsx"
    monkeypatch.setattr(
        writeback._StyleCloner,
        "styled",
        lambda *args: (_ for _ in ()).throw(ValueError("bad style")),
    )

    report = writeback.write(model, str(output))

    assert report.warnings == (
        "Style not applied to B3: ValueError: bad style",
        "Style not applied to C3: ValueError: bad style",
    )
    with zipfile.ZipFile(output) as package:
        root = etree.fromstring(package.read("xl/worksheets/sheet1.xml"))
    assert root.xpath(
        "string(.//m:c[@r='B3']/m:is/m:t)", namespaces={"m": writeback.NS_MAIN}
    ) == "Approved "


def test_style_failure_keeps_untouched_styles_byte_for_byte(monkeypatch, tmp_path: Path) -> None:
    source, model = _crosswalk(tmp_path)
    output = tmp_path / "out.xlsx"
    monkeypatch.setattr(
        writeback._StyleCloner,
        "styled",
        lambda *args: (_ for _ in ()).throw(ValueError("bad style")),
    )

    writeback.write(model, str(output))

    with zipfile.ZipFile(source) as source_zip, zipfile.ZipFile(output) as output_zip:
        assert source_zip.read("xl/styles.xml") == output_zip.read("xl/styles.xml")


def test_style_cloner_construction_failure_is_reported(monkeypatch, tmp_path: Path) -> None:
    _, model = _crosswalk(tmp_path)
    output = tmp_path / "out.xlsx"

    monkeypatch.setattr(
        writeback,
        "_StyleCloner",
        lambda *_: (_ for _ in ()).throw(ValueError("styles unavailable")),
    )

    report = writeback.write(model, str(output))

    assert report.warnings == ("Style cloning disabled: ValueError: styles unavailable",)


def test_verify_rejects_changed_package_members(tmp_path: Path) -> None:
    source, model = _crosswalk(tmp_path)
    output = tmp_path / "out.xlsx"
    writeback.write(model, str(output))
    with zipfile.ZipFile(output, "a") as package:
        package.writestr("unexpected.txt", b"not in source")

    with pytest.raises(OutputError, match="package members"):
        writeback.verify_output(
            source,
            output,
            (writeback.CellEdit("xl/worksheets/sheet1.xml", "B3", "Approved "),),
            styles_changed=True,
        )


def test_verify_rejects_corrupted_hidden_sheet_and_lookup_spec(tmp_path: Path) -> None:
    source, model = _crosswalk(tmp_path)
    output = tmp_path / "out.xlsx"
    writeback.write(model, str(output))
    _rewrite_package(
        output,
        lambda name, content: content + b"corrupt" if name == "xl/worksheets/sheet2.xml" else content,
    )

    with pytest.raises(OutputError, match="Untouched workbook part"):
        writeback.verify_output(
            source,
            output,
            (writeback.CellEdit("xl/worksheets/sheet1.xml", "B3", "Approved "),),
            styles_changed=True,
        )

    writeback.write(model, str(output), overwrite=True)
    _rewrite_package(
        output,
        lambda name, content: content + b"corrupt" if name == "xl/worksheets/sheet3.xml" else content,
    )

    with pytest.raises(OutputError, match="Untouched workbook part"):
        writeback.verify_output(
            source,
            output,
            (writeback.CellEdit("xl/worksheets/sheet1.xml", "B3", "Approved "),),
            styles_changed=True,
        )


def test_verify_rejects_missing_hidden_sheet_and_lookup_spec(tmp_path: Path) -> None:
    source, model = _crosswalk(tmp_path)
    output = tmp_path / "out.xlsx"
    expected = (writeback.CellEdit("xl/worksheets/sheet1.xml", "B3", "Approved "),)
    writeback.write(model, str(output))
    _rewrite_package(output, lambda name, content: None if name == "xl/worksheets/sheet2.xml" else content)

    with pytest.raises(OutputError, match="package members"):
        writeback.verify_output(source, output, expected, styles_changed=True)

    writeback.write(model, str(output), overwrite=True)
    _rewrite_package(output, lambda name, content: None if name == "xl/worksheets/sheet3.xml" else content)

    with pytest.raises(OutputError, match="package members"):
        writeback.verify_output(source, output, expected, styles_changed=True)


def test_verify_requires_inline_string_cells_with_preserved_space(tmp_path: Path) -> None:
    source, model = _crosswalk(tmp_path)
    output = tmp_path / "out.xlsx"
    writeback.write(model, str(output))

    def replace_cell(name: str, content: bytes) -> bytes:
        if name != "xl/worksheets/sheet1.xml":
            return content
        root = etree.fromstring(content)
        cell = root.xpath(".//m:c[@r='B3']", namespaces={"m": writeback.NS_MAIN})[0]
        cell.attrib.pop("t")
        for child in list(cell):
            cell.remove(child)
        value = etree.SubElement(cell, f"{{{writeback.NS_MAIN}}}v")
        value.text = "Approved "
        return etree.tostring(root, xml_declaration=True, encoding="UTF-8")

    _rewrite_package(output, replace_cell)

    with pytest.raises(OutputError, match="inline string"):
        writeback.verify_output(
            source,
            output,
            (writeback.CellEdit("xl/worksheets/sheet1.xml", "B3", "Approved "),),
            styles_changed=True,
        )
