import json
from pathlib import Path

import openpyxl
import pytest

from conversion_agent.core.errors import WorkbookError
from conversion_agent.mapping import service
from conversion_agent.mapping.validation import load_validated_workbook
from conversion_agent.mapping.validation import validate_proposal_document


def test_rejects_duplicate_proposal_keys() -> None:
    payload = {
        "proposals": [
            {
                "tab": "Permits",
                "section": "Type",
                "source": ["A"],
                "dest": ["One"],
                "confidence": 0.8,
                "rationale": "first",
            },
            {
                "tab": "Permits",
                "section": "Type",
                "source": ["A"],
                "dest": ["Two"],
                "confidence": 0.7,
                "rationale": "second",
            },
        ]
    }
    with pytest.raises(WorkbookError, match="[Dd]uplicate"):
        validate_proposal_document(payload)


def test_rejects_workbook_without_lookup_spec(tmp_path) -> None:
    path = tmp_path / "invalid.xlsx"
    openpyxl.Workbook().save(path)
    with pytest.raises(WorkbookError, match="LookupSpec"):
        load_validated_workbook(path)


def test_wraps_a_corrupt_workbook_as_a_workbook_error(tmp_path) -> None:
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"not an XLSX archive")

    with pytest.raises(WorkbookError, match="Invalid workbook"):
        load_validated_workbook(path)


def _write_lookup_spec(workbook, payload: object) -> None:
    sheet = workbook.active
    sheet.title = "LookupSpec"
    sheet.cell(row=1, column=1, value="spec")
    sheet.cell(row=1, column=5, value=json.dumps(payload))


def test_rejects_lookup_spec_with_a_non_object_json_shape(tmp_path) -> None:
    path = tmp_path / "list-spec.xlsx"
    workbook = openpyxl.Workbook()
    _write_lookup_spec(workbook, [])
    workbook.save(path)

    with pytest.raises(WorkbookError, match="LookupSpec"):
        load_validated_workbook(path)


def test_rejects_short_hidden_sheet_rows_as_workbook_errors(tmp_path) -> None:
    path = tmp_path / "short-hidden-row.xlsx"
    workbook = openpyxl.Workbook()
    _write_lookup_spec(workbook, {"modules": {}})
    workbook.create_sheet("Permits Hidden").append(["short row"])
    workbook.save(path)

    with pytest.raises(WorkbookError, match="hidden"):
        load_validated_workbook(path)


def test_rejects_cascade_children_before_deterministic_matching(monkeypatch, tmp_path) -> None:
    path = tmp_path / "invalid-cascade-child.xlsx"
    workbook = openpyxl.Workbook()
    _write_lookup_spec(
        workbook,
        {
            "modules": {
                "Permits": {
                    "typeQueries": {
                        "Type": {
                            "source": {"columns": ["SOURCE"]},
                            "destination": {"columns": ["PARENT", "CHILD"]},
                        }
                    }
                }
            }
        },
    )
    permits = workbook.create_sheet("Permits")
    permits.append(["Type"])
    permits.append(["Source DB", "Destination DB", "Destination DB"])
    permits.append(["Source", "Parent", "Child"])
    permits.append(["Parent", "", ""])
    hidden = workbook.create_sheet("Permits Hidden")
    hidden.append(["Type", 1, 2])
    hidden.append(["Type.head", None, None, "Parent"])
    hidden.append(["Type.head", None, None, "Configured child"])
    hidden.append(["Parent", None, 1, "Unconfigured child"])
    workbook.save(path)

    matcher_called = False

    def cannot_match(*args, **kwargs) -> None:
        nonlocal matcher_called
        matcher_called = True

    monkeypatch.setattr(service.match, "run", cannot_match)

    with pytest.raises(WorkbookError, match="cascade child"):
        service.MappingService().run(
            service.MappingRequest(input_path=path, output_path=Path(tmp_path / "out.xlsx"))
        )

    assert not matcher_called


@pytest.mark.parametrize(
    "payload",
    [
        {
            "proposals": [
                {
                    "tab": "Permits",
                    "section": "Type",
                    "source": ["A"],
                    "dest": ["One"],
                    "confidence": 1.1,
                }
            ]
        },
        {
            "proposals": [
                {
                    "tab": "Permits",
                    "section": "Type",
                    "source": ["A"],
                    "dest": ["One"],
                    "unexpected": True,
                }
            ]
        },
    ],
)
def test_rejects_invalid_proposal_schema(payload: object) -> None:
    with pytest.raises(WorkbookError, match="Invalid proposal document"):
        validate_proposal_document(payload)
